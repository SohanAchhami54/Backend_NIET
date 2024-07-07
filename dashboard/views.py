import json
import pandas as pd
from datetime import datetime
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.utils.timezone import now
import openpyxl
from openpyxl.utils.datetime import from_excel
from django.shortcuts import get_object_or_404


# from general.models import UserType, AppUser, StudentBatch, Student
from django.contrib.auth.hashers import make_password
import random
import string

from dashboard.models import *
from general.models import *
from dashboard.serializers import *
from django.http import Http404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser





from celery import shared_task

from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags

# Create your views here.
# Create your views here.
def admin_dashboard(request):
    students = Student.objects.filter(is_active=True,is_alumni=True)
    student_count =  Student.objects.filter(is_active=True,is_alumni=True).count()
    visitingfaculty_count =  Faculty.objects.filter(
        faculty_type__type_name='visiting', 
        faculty_type__is_active=True
    ).count()
    fullfaculty_count =  Faculty.objects.filter(
        faculty_type__type_name='full time', 
        faculty_type__is_active=True
    ).count()
    context_dict = {
        'students':students,'student_count':student_count,
        'visitingfaculty_count':visitingfaculty_count,
        'fullfaculty_count':fullfaculty_count,
        }
    return render(request,'dashboard/index.html',context_dict)

@shared_task
def send_registration_email(user, password):
    subject = 'Welcome to Our College!'
    context = {
        'user': user,
        'password': password,
        'year': datetime.now().year
    }
    html_message = render_to_string('email_template.html', context)
    plain_message = strip_tags(html_message)
    from_email = 'infymee@gmail.com'
    to = user.email
    
    send_mail(subject, plain_message, from_email, [to], html_message=html_message)

def dashboard(request):
    return render(request, "index.html")




def generate_random_password(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for i in range(length))

def bulk_upload_students(request):
    if request.method == "POST":
        try:
            
            excel_file = request.FILES["file"]
            
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            user_type = UserType.objects.get(name="Student")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(row)
                first_name, middle_name, last_name, email, batch_name, admission_year, passed_year, registration_number = row
                if not email:
                    continue  # Skip rows with no email
                
                # Create AppUser
                password = generate_random_password()
                
                user = AppUser.objects.filter(email=email)
                if user:
                    continue
                
                user = AppUser.objects.create(
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    email=email,
                    is_active=True,
                    is_staff=False,
                    date_joined=now(),
                    password=make_password(password)
                )
                print(first_name,password)

                admission_year = str(admission_year).split(".")[0]
                passed_year = str(passed_year).split(".")[0]

                admission_year = datetime.strptime(admission_year,"%Y")
                passed_year  = datetime.strptime(passed_year,"%Y")

                
                print(admission_year,passed_year)
                
                # Get or create StudentBatch

                # batch, created = StudentBatch.objects.get_or_create(
                #     name=batch_name,
                #     admission_year=admission_year,
                #     passed_year=passed_year,
                #     defaults={'is_active': True}
                # )

                # Create Student
                student = Student.objects.create(
                    user=user,
                    user_type=user_type,
                    # batch=batch,
                    registration_number=registration_number
                    # is_alumni = is_alumni
                )
                try:
                    send_registration_email(user,password)
                except Exception as e:
                    print(e)
            
            return JsonResponse({'message': 'Students have been uploaded successfully.'}, status=200)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=400)
    return render(request, "dashboard/index.html")


def bulk_upload_faculty(request):
    if request.method == "POST":
        try:
            
            excel_file = request.FILES["file"]
            
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            user_type = UserType.objects.get(name="Teacher")
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                first_name, middle_name, last_name, email, education,designation,post_name,faculty_type = row
                
                if not (first_name and last_name and email and education and designation and post_name and faculty_type):
                    continue  
                if not email:
                    continue  # Skip rows with no email
                
                # Create AppUser
                password = generate_random_password()
                
                user = AppUser.objects.filter(email=email)
                if user:
                    continue
                
                user = AppUser.objects.create(
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    email=email,
                    is_active=True,
                    is_staff=False,
                    date_joined=now(),
                    password=make_password(password)
                )
                
                ftype,created = FacultyType.objects.get_or_create(type_name=faculty_type)

                # Create Faculty
                Faculty.objects.create(
                    user=user,
                    user_type=user_type,
                    education = education,
                    designation = designation,
                    post_name = post_name,
                    faculty_type = ftype
                )

            
            return JsonResponse({'message': 'Faculty have been uploaded successfully.'}, status=200)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=400)
    return render(request, "index.html")

# prakash Z9Iqwo6c




# API view defined here

class ScheduleList(APIView):
    """
    List all yearlyschedule, or create a new yearlyschedule.
    """
    def get(self, request, format=None):
        items = Schedule.objects.filter(is_active=True)
        serializer = ScheduleSerializer(items, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = ScheduleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class YearlySchedulerList(APIView):
    """
    List all yearlyschedule, or create a new yearlyschedule.
    """
    def get(self, request, format=None):
        items = YearlyScheduler.objects.filter(is_active=True).order_by('-created_at')
        serializer = YearlySchedulerSerializerDetail(items, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = YearlySchedulerSerializer(data=request.data)
        print(serializer)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class StudentTheoryInternalMarkUpload(APIView):
    parser_classes = (MultiPartParser,FormParser)
    def post(self,request):
        file = request.data.get('file')
        subject_id = request.data.get('subject')
        schedule_id = request.data.get('schedule')
        if not file or not subject_id or not schedule_id:
            return Response(
                {"error": "File, yearly schedule and subject are required"},
                status=status.HTTP_400_BAD_REQUEST
            )        
        try:
            df = pd.read_excel(file)
            for _,row in df.iterrows():
                subject = Subject.objects.get(id=subject_id)
                student = Student.objects.get(registration_number=row['registration_number'])
                yearly_scheduler = YearlyScheduler.objects.get(id=schedule_id)
                subject_schedule = SubjectSchedule.objects.get(subject=subject,schedule=yearly_scheduler.schedule)
                student_yearly_scheduler = StudentYearlyScheduler.objects.get(student=student,yearly_schedule = yearly_scheduler)
                im = TheoryInternalMarks.objects.filter(student_yearly_scheduler=student_yearly_scheduler,subject = subject_schedule)

                if(im):
                    return Response(
                        {"error": "ALready uploaded"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                internalmark = TheoryInternalMarks.objects.create(
                    student_yearly_scheduler=student_yearly_scheduler,
                    subject = subject_schedule,
                    internal_assessment = row['internal_assessment'],
                    attendance = row['attendance'],
                    class_performance = row['class_performance'],
                    assignment = row['assignment'],
                    presentation = row['presentation'],
                    remarks = row['remarks']
                    )
            return Response(
                {"message": "Students uploaded and linked to the yearly schedule successfully"},
                status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class StudentPracticalInternalMarkUpload(APIView):
    parser_classes = (MultiPartParser,FormParser)
    def post(self,request):
        file = request.data.get('file')
        subject_id = request.data.get('subject')
        schedule_id = request.data.get('schedule')
        if not file or not subject_id or not schedule_id:
            return Response(
                {"error": "File, yearly schedule and subject are required"},
                status=status.HTTP_400_BAD_REQUEST
            )        
        try:
            df = pd.read_excel(file)
            for _,row in df.iterrows():
                subject = Subject.objects.get(id=subject_id)
                student = Student.objects.get(registration_number=row['registration_number'])
                yearly_scheduler = YearlyScheduler.objects.get(id=schedule_id)
                subject_schedule = SubjectSchedule.objects.get(subject=subject,schedule=yearly_scheduler.schedule)
                student_yearly_scheduler = StudentYearlyScheduler.objects.get(student=student,yearly_schedule = yearly_scheduler)
                im = PracticalInternalMarks.objects.filter(student_yearly_scheduler=student_yearly_scheduler,subject = subject_schedule)
                if(im):
                    return Response(
                        {"error": "ALready uploaded"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )


                internalmark = PracticalInternalMarks.objects.create(
                    student_yearly_scheduler=student_yearly_scheduler,
                    subject = subject_schedule,
                    attendance = row['attendance'],
                    labexam_viva = row['labexam_viva'],
                    lab_report = row['lab_report']
                )
            return Response(
                {"message": "Students uploaded and linked to the yearly schedule successfully"},
                status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class StudentYearlyScheduleUpload(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file = request.data.get('file')
        yearly_schedule_id = request.data.get('yearly_schedule')

        if not file or not yearly_schedule_id:
            return Response(
                {"error": "File and yearly schedule are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            yearly_schedule = YearlyScheduler.objects.get(id=yearly_schedule_id)
        except YearlyScheduler.DoesNotExist:
            return Response(
                {"error": "Yearly schedule not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            df = pd.read_excel(file)
            # students = []
            user_type = UserType.objects.get(name="Student")

            for _, row in df.iterrows():
                if not row['email']:
                    continue  
                
                # Create AppUser
                password = generate_random_password()
                
                user = AppUser.objects.filter(email=row['email'])
                if user:
                    continue
                
                user = AppUser.objects.create(
                    first_name=row['first_name'],
                    middle_name=row['middle_name'],
                    last_name=row['last_name'],
                    email=row['email'],
                    is_active=True,
                    is_staff=False,
                    date_joined=now(),
                    password=make_password(password)
                )
                student, created = Student.objects.get_or_create(
                    user=user,
                    user_type=user_type,
                    # batch=batch,
                    registration_number=row['registration_number']
                    # is_alumni = is_alumni
                )
                # students.append(student)
                student_yearly_scheduler = StudentYearlyScheduler.objects.create(
                    yearly_schedule=yearly_schedule,
                    student=student
                    )
                student_yearly_scheduler.save()

            return Response(
                {"message": "Students uploaded and linked to the yearly schedule successfully"},
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

class StudentYearlyScheduleList(APIView):
    def get(self, request,schedule_id):
        yearly_schedule_id = schedule_id
        try:
            yearly_schedule = StudentYearlyScheduler.objects.filter(yearly_schedule__id=yearly_schedule_id,is_active=True)
            serializer = StudentYearlySchedulerSerializer(yearly_schedule,many=True)
            return Response(
                serializer.data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class StudentYearlyScheduleUpdate(APIView):
    def post(self,request):
        yearly_schedule_id = request.data.get('yearly_schedule')
        prev_schedule_id = request.data.get('previous_schedule')
        student_ids = request.data.get('selected_ids').split(',')
        yearly_schedule = YearlyScheduler.objects.get(id=yearly_schedule_id)
        prev_yearly_schedule = YearlyScheduler.objects.get(id=prev_schedule_id)
        for id in student_ids:
            student = Student.objects.get(id=id)
            prev_student_yearly_schedule = StudentYearlyScheduler.objects.get(yearly_schedule=prev_yearly_schedule,student=student)
            prev_student_yearly_schedule.is_active=False
            prev_student_yearly_schedule.save()
            student_yearly_schedule = StudentYearlyScheduler.objects.create(yearly_schedule=yearly_schedule,student=student)
            student_yearly_schedule.is_active = True
            student_yearly_schedule.save()
        return Response(
            {"message": "Students yearly schedule updated successfully"},
            status=status.HTTP_202_ACCEPTED
        )
    

class SubjectList(APIView):
    """
    List all Subject, or create a new Subject.
    """
    def get(self, request, format=None):
        items = Subject.objects.filter(is_active=True)
        serializer = SubjectSerializer(items, many=True)
        return Response(serializer.data)

    def post(self, request, format=None):
        serializer = SubjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SubjectListBySchedule(APIView):
    """
    List all subject by its schedule year
    """
    def get(self,request,schedule_year_id):
        try:
            schedule_year = YearlyScheduler.objects.get(id=schedule_year_id)
            schedule = Schedule.objects.get(id=schedule_year.schedule.id)
            subjects = SubjectSchedule.objects.filter(schedule__id=schedule.id)
        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        serializers = SubjectScheduleSerializer(subjects,many=True)
        return Response(serializers.data, status=status.HTTP_200_OK)
    

class StudentInternalMarksList(APIView):
    def get(self, request, student_yearly_schedule_id):
        try:
            student_scheduler = StudentYearlyScheduler.objects.get(id=student_yearly_schedule_id)
        except StudentYearlyScheduler.DoesNotExist:
            return Response({"error": "Student Yearly Scheduler does not exist"}, status=status.HTTP_404_NOT_FOUND)
        
        theorymarks = TheoryInternalMarks.objects.filter(student_yearly_scheduler=student_scheduler)
        theory_serializers = TheoryInternalMarksSerializer(theorymarks, many=True)
        
        practicalmarks = PracticalInternalMarks.objects.filter(student_yearly_scheduler=student_scheduler)
        practical_serializers = PracticalInternalMarksSerializer(practicalmarks, many=True)
        
        return Response({
            "theory_marks": theory_serializers.data,
            "practical_marks": practical_serializers.data
        })


class StudentCountBySchedule(APIView):
    def get(self, request, yearly_schedule_id):
        yearly_scheduler = get_object_or_404(YearlyScheduler, id=yearly_schedule_id)
        student_count = StudentYearlyScheduler.objects.filter(yearly_schedule=yearly_scheduler).count()
        return Response({'yearly_schedule_id': yearly_schedule_id, 'student_count': student_count}, status=status.HTTP_200_OK)





    





